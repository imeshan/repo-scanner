import sys
from openpyxl import load_workbook

input_file = sys.argv[1]

workbook = load_workbook(input_file)
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    repo_url, package, required_version = row

    print(f"Repository: {repo_url}")
    print(f"Package: {package}")
    print(f"Required version: {required_version}")
    print("-" * 40)
